"""Importação segura de portfólio: somente preview, sem persistência."""
from __future__ import annotations

import csv
import io
import json
import re
import unicodedata

from app.schemas.intake import DppIntakePayload, IntakeImportResult, IntakeRowResult
from app.services.intake_validator import validate_intake

MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_ROWS = 5000

ALIASES = {
    "id produto": "product_id", "produto id": "product_id", "product id": "product_id",
    "nome produto": "product_name", "produto": "product_name", "product name": "product_name",
    "descricao": "product_description", "descrição": "product_description",
    "categoria": "product_category", "marca": "brand_name", "brand": "brand_name",
    "codigo barras": "gtin", "código barras": "gtin", "lote": "batch_id",
    "pais fabricacao": "country_of_manufacture", "país fabricação": "country_of_manufacture",
    "composicao": "fiber_composition", "composição": "fiber_composition",
}


def _key(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


CANONICAL = {_key(field): field for field in DppIntakePayload.model_fields}
NORMALIZED_ALIASES = {_key(key): value for key, value in ALIASES.items()}


def _decode_cell(field: str, value):
    if value is None or value == "":
        return None
    if field == "fiber_composition" and isinstance(value, str):
        parts = []
        for item in re.split(r"[;|]", value):
            match = re.match(r"\s*(.+?)\s*[:=]\s*([0-9.,]+)\s*%?\s*$", item)
            if match:
                parts.append({"fibra": match.group(1), "pct": float(match.group(2).replace(",", "."))})
        return parts or value
    if isinstance(value, str) and value.strip().startswith(("[", "{")):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value


def _map_rows(headers, raw_rows):
    mapped, unmapped = {}, []
    for header in headers:
        normalized = _key(str(header))
        target = CANONICAL.get(normalized) or NORMALIZED_ALIASES.get(normalized)
        if target:
            mapped[str(header)] = target
        else:
            unmapped.append(str(header))
    rows = [{target: _decode_cell(target, row.get(source)) for source, target in mapped.items()} for row in raw_rows]
    return mapped, unmapped, rows


def parse_portfolio(filename: str, content: bytes) -> IntakeImportResult:
    if len(content) > MAX_FILE_BYTES:
        raise ValueError("arquivo excede 10 MB")
    extension = filename.rsplit(".", 1)[-1].lower()
    if extension == "csv":
        text = content.decode("utf-8-sig")
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        raw_rows, headers = list(reader), reader.fieldnames or []
    elif extension == "xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(values, ())]
        raw_rows = [dict(zip(headers, row)) for row in values]
    elif extension == "json":
        decoded = json.loads(content.decode("utf-8"))
        raw_rows = decoded if isinstance(decoded, list) else decoded.get("products", [])
        headers = list(dict.fromkeys(key for row in raw_rows for key in row))
    else:
        raise ValueError("formato aceito: CSV, XLSX ou JSON")
    if len(raw_rows) > MAX_ROWS:
        raise ValueError("arquivo excede 5.000 produtos")
    mapped, unmapped, rows = _map_rows(headers, raw_rows)
    results = []
    for number, row in enumerate(rows, 2):
        validation = validate_intake(row)
        results.append(IntakeRowResult(row_number=number, product_reference=row.get("sku") or row.get("product_id"), validation=validation))
    valid = sum(item.validation.valid for item in results)
    return IntakeImportResult(filename=filename, format=extension, total_rows=len(results), valid_rows=valid,
        invalid_rows=len(results)-valid, mapped_columns=mapped, unmapped_columns=unmapped, rows=results)
